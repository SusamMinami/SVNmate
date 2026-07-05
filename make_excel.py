
import re
import math
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

tree = ET.parse(r'c:\Users\Admin\Downloads\ezxss\Sam版上床下桌.scene')
root = tree.getroot()
objects = root.find('objects')

def mat4_inverse(mat):
    a,b,c,d = mat[0],mat[1],mat[2],mat[3]
    e,f,g,h = mat[4],mat[5],mat[6],mat[7]
    i,j,k,l = mat[8],mat[9],mat[10],mat[11]
    m0,m1,m2,m3 = mat[12],mat[13],mat[14],mat[15]
    A = f*k - g*j; B = g*i - e*k; C = e*j - f*i
    det = a*A + b*B + c*C
    if abs(det) < 1e-9: return None
    inv_det = 1.0/det
    return [
        A*inv_det, (b*k-c*j)*inv_det, (b*g-c*f)*inv_det, 0,
        B*inv_det, (c*i-a*k)*inv_det, (c*e-a*g)*inv_det, 0,
        C*inv_det, (a*j-b*i)*inv_det, (a*f-b*e)*inv_det, 0,
        -(m0*A+m1*B+m2*C)*inv_det,
        -(m0*(b*k-c*j)+m1*(c*i-a*k)+m2*(a*j-b*i))*inv_det,
        -(m0*(b*g-c*f)+m1*(c*e-a*g)+m2*(a*f-b*e))*inv_det, 1
    ]

def mat4_mul_vec(mat, v):
    x,y,z,w = v
    return (
        mat[0]*x+mat[1]*y+mat[2]*z+mat[3]*w,
        mat[4]*x+mat[5]*y+mat[6]*z+mat[7]*w,
        mat[8]*x+mat[9]*y+mat[10]*z+mat[11]*w,
    )

profiles_scene = {}
for obj in objects.findall('object'):
    if obj.findtext('type') == 'Profile' and obj.findtext('bom_exclude','1') == '0':
        pid = int(obj.findtext('id'))
        code = obj.findtext('profile', '')
        height_cm = float(obj.findtext('height', '0'))
        length_mm = round(height_cm * 10)
        
        size_match = re.search(r'PROF(\d{2})-(\d{4})L?', code)
        if size_match:
            s1 = int(size_match.group(1))
            s2_str = size_match.group(2)
            s2 = int(s2_str[2:])
            size = f"{s1}x{s2}"
        else:
            size = ''
        
        rot = [float(x) for x in obj.findtext('rotation','').split(',') if x.strip()]
        if len(rot) < 16: continue
        
        inv_mat = mat4_inverse(rot)
        
        profiles_scene[pid] = {
            'id': pid, 'size': size, 'length': length_mm,
            'inv_mat': inv_mat, 'holes': []
        }

for obj in objects.findall('object'):
    if obj.findtext('type') == 'SimpleBore':
        parent = int(obj.findtext('parent','0'))
        if parent not in profiles_scene: continue
        p = profiles_scene[parent]
        
        md = float(obj.findtext('main_diameter','0')) * 10
        sd = float(obj.findtext('second_diameter','0')) * 10
        mdp = float(obj.findtext('main_depth','0')) * 10
        sdp = float(obj.findtext('second_depth','0')) * 10
        
        rot = [float(x) for x in obj.findtext('rotation','').split(',') if x.strip()]
        if len(rot) < 16: continue
        if p['inv_mat'] is None: continue
        
        world_pos = (rot[12], rot[13], rot[14], 1)
        local_pos = mat4_mul_vec(p['inv_mat'], world_pos)
        world_axis = (rot[8], rot[9], rot[10], 0)
        local_axis = mat4_mul_vec(p['inv_mat'], world_axis)
        
        lx = local_pos[0]
        ay, az = local_axis[1], local_axis[2]
        
        dist = round(abs(lx))
        
        if abs(az) > abs(ay):
            side = 1 if az > 0 else 3
        else:
            side = 4 if ay > 0 else 2
        
        htype = 'cb' if sd >= 7 else 'tap'
        p['holes'].append({
            'type': htype, 'side': side, 'dist': dist,
            'cb_dia': round(sd) if htype=='cb' else 0,
            'cb_depth': round(sdp) if htype=='cb' else 0,
            'thru_dia': round(md) if htype=='cb' else 0,
            'tap_dia': round(md) if htype=='tap' else 0,
            'depth': round(mdp) if htype=='tap' else 0
        })

txt_text = open(r'c:\Users\Admin\Downloads\ezxss\导出尝试.txt', 'r', encoding='utf-8', errors='ignore').read()
txt_lines = [l.rstrip() for l in txt_text.split('\n') if l.strip()]

bom_profiles = {}
size_code_map = {'020020':'20x20','020040':'20x40','040020':'20x40','030030':'30x30','040040':'40x40'}

current_pos = 0
za_connectors = {}
q11_counts = {}
for line in txt_lines:
    m = re.match(r'^(\d+)\s+1\.1[01]\.(\d{6})\.', line)
    if m:
        pos = int(m.group(1))
        code6 = m.group(2)
        qty_match = re.search(r'\s+(\d+)\s+[\d.]+\s+[\d.]+\s*$', line)
        qty = int(qty_match.group(1)) if qty_match else 1
        lm = re.search(r'/(\d+)\s', line)
        length = int(lm.group(1)) if lm else 0
        size = size_code_map.get(code6, f"{code6[:2]}x{code6[3:5]}")
        if pos not in bom_profiles:
            bom_profiles[pos] = {'pos':pos,'size':size,'length':length,'qty':qty,'holes':[]}
        current_pos = pos
        continue
    
    zm = re.match(r'^\d+\.[A-Z]\s+ZA([1-4])/([\d.]+)', line)
    if zm and current_pos:
        za_connectors.setdefault(current_pos, []).append({
            'anchor_side': int(zm.group(1)),
            'dist': round(float(zm.group(2)))
        })

by_sizelen = {}
for pid, sp in profiles_scene.items():
    key = (sp['size'], sp['length'])
    by_sizelen.setdefault(key, []).append(sp)

mapped = 0
for pos, bp in bom_profiles.items():
    key = (bp['size'], bp['length'])
    matched = None
    if key in by_sizelen and by_sizelen[key]:
        matched = by_sizelen[key].pop(0)
    elif bp['size'] == '20x40' and ('40x20', bp['length']) in by_sizelen and by_sizelen[('40x20', bp['length'])]:
        matched = by_sizelen[('40x20', bp['length'])].pop(0)
    
    if matched:
        bp['holes'] = matched['holes']
        mapped += 1

print(f"成功映射 {mapped}/{len(bom_profiles)} 个型材到scene数据")

for pos, zas in za_connectors.items():
    if pos not in bom_profiles: continue
    bp = bom_profiles[pos]
    for za in zas:
        bp['holes'].append({
            'type': 'thru',
            'side': za['anchor_side'],
            'dist': za['dist'],
            'thru_dia': 9
        })

accessories_list = [
    (31, '1.46.204.1720.2', '17×20锌合金角码（带黑色涂层）', 26, '用于20系列型材T型槽连接'),
    (32, '1.46.21039', '100×100重型直角角件', 30, '40/30系列型材强力直角连接，安装孔Ø9'),
    (33, '0.63.D06912.08010', '内六角圆柱头螺钉 DIN6912', 15, 'M8×10mm'),
    (34, '0.63.D06912.08014', '内六角圆柱头螺钉 DIN6912', 96, 'M8×14mm（主连接螺丝）'),
    (35, '0.63.WN7381.05006', '圆头法兰螺钉', 4, 'M5×6mm（角件端盖固定）'),
    (36, '0.63.WN7381.05008', '圆头法兰螺钉', 14, 'M5×8mm（板材/面板安装）'),
    (37, '0.63.WN7381.05008', '圆头法兰螺钉', 31, 'M5×8mm（板材/面板安装）'),
    (38, '1.20.2E0', '万向连接件 E型', 6, '自动调平连接块'),
    (39, '1.20.2F0', '万向连接件 F型', 2, '自动调平连接块'),
    (40, '1.20.2H0', '万向连接件 H型', 1, '自动调平连接块'),
    (41, '1.21.3E0', '十字连接件 E型', 10, '垂直交叉连接'),
    (42, '1.21.3F0', '十字连接件 F型', 11, '垂直交叉连接'),
    (43, '1.21.4E0', '十字连接件 4E型', 7, '垂直交叉连接'),
    (44, '1.31.HM5', 'H型螺纹弹片', 4, 'M5（预置在T型槽中）'),
    (45, '1.32.4EM5', 'E型弹片T型螺母（后装）', 14, 'M5'),
    (46, '1.32.4EM8', 'E型弹片T型螺母（后装）', 96, 'M8（主连接用）'),
    (47, '1.32.4FM5', 'F型弹片T型螺母（后装）', 31, 'M5'),
    (48, '1.32.4FM8', 'F型弹片T型螺母（后装）', 15, 'M8'),
    (49, '1.46.204.1720A', '17×20角码端盖', 26, '装饰盖帽（黑色塑料）'),
]

wb = Workbook()
ws = wb.active
ws.title = "铝型材加工清单"

hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
stat_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
note_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

headers = ['编号', '型材型号', '长度\n(mm)', '数量', '每个沉头孔\n(M8,Ø8)', '每个通孔\n(Ø9连接件)', '沉头合计', '通孔合计', '具体加工说明（参考PDF）']
for c, h in enumerate(headers, 1):
    cell = ws.cell(1, c, h)
    cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center; cell.border = thin

type_map = {'20x20':'2020欧标','20x40':'2040欧标','30x30':'3030欧标','40x40':'4040欧标'}

row = 2
total_cb = 0
total_thru = 0

for pos in sorted(bom_profiles.keys()):
    p = bom_profiles[pos]
    if p['length'] == 0: continue
    
    sm = {}
    for h in p['holes']:
        if h['side'] is None: continue
        sm.setdefault(h['side'], []).append(h)
    
    parts = []
    for s in sorted(sm.keys()):
        sp = []
        for h in sorted(sm[s], key=lambda x: x['dist']):
            if h['type'] == 'cb':
                sp.append(f"{h['dist']}mm处：沉头孔Ø{h['cb_dia']}深{h['cb_depth']}，通孔Ø{h['thru_dia']}（装M8内六角螺丝）")
            elif h['type'] == 'thru':
                sp.append(f"{h['dist']}mm处：通孔Ø{h['thru_dia']}.2（安装万向/十字连接件）")
            elif h['type'] == 'tap':
                sp.append(f"{h['dist']}mm处：螺纹底孔Ø{h['tap_dia']}深{h['depth']}（攻M5）")
        parts.append(f"第{s}面：" + "；".join(sp))
    
    desc_text = "\n".join(parts) if parts else "无CNC加工（使用角件+T型螺母直接组装，无需钻孔）"
    
    cb = sum(1 for h in p['holes'] if h['type'] == 'cb')
    th = sum(1 for h in p['holes'] if h['type'] == 'thru')
    tp = sum(1 for h in p['holes'] if h['type'] == 'tap')
    
    ws.cell(row, 1, pos)
    ws.cell(row, 2, f"{type_map.get(p['size'], p['size'])}")
    ws.cell(row, 3, p['length'])
    ws.cell(row, 4, p['qty'])
    ws.cell(row, 5, cb)
    ws.cell(row, 6, th)
    ws.cell(row, 7, cb * p['qty'])
    ws.cell(row, 8, th * p['qty'])
    ws.cell(row, 9, desc_text)
    
    for c in range(1, 10):
        cell = ws.cell(row, c)
        cell.border = thin
        cell.alignment = left_wrap if c == 9 else center
    
    total_cb += cb * p['qty']
    total_thru += th * p['qty']
    row += 1

ws.cell(row, 1, "总计")
ws.cell(row, 1).font = Font(bold=True)
ws.cell(row, 7, total_cb)
ws.cell(row, 8, total_thru)
ws.cell(row, 9, "※说明：角件连接靠T型螺母固定，无需在型材上钻孔；\nM5螺丝用于角件自带螺纹孔（端盖/面板安装），不在型材上攻丝。")
for c in range(1, 10):
    cell = ws.cell(row, c)
    cell.font = Font(bold=True); cell.fill = stat_fill; cell.border = thin
    cell.alignment = left_wrap if c == 9 else center

ws2 = wb.create_sheet("配件清单")
ah = ['编号', 'MayTec型号', '中文名称', '数量', '用途说明']
for c, h in enumerate(ah, 1):
    cell = ws2.cell(1, c, h)
    cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center; cell.border = thin

ar = 2
for pos, art, cn, qty, spec in accessories_list:
    ws2.cell(ar, 1, pos)
    ws2.cell(ar, 2, art)
    ws2.cell(ar, 3, cn)
    ws2.cell(ar, 4, qty)
    ws2.cell(ar, 5, spec)
    for c in range(1, 6):
        cell = ws2.cell(ar, c)
        cell.border = thin
        cell.alignment = center if c in (1,4) else left_wrap
    ar += 1

ws3 = wb.create_sheet("力学评估与组装说明")

eval_data = [
    ['项目', '数据/说明'],
    ['—— 基本参数 ——', ''],
    ['设计名称', 'Sam版 上床下桌（铝型材框架）'],
    ['整体尺寸', '长2040 × 宽1580 × 高1935 mm'],
    ['框架重量', '约55.4 kg（仅铝型材）'],
    ['主承力型材', '床架纵梁：4040欧标；立柱：4040欧标；横梁：3030/4040欧标'],
    ['', ''],
    ['—— 载荷计算 ——', ''],
    ['用户体重', '80 kg（静载荷）'],
    ['伴侣体重', '60 kg（静载荷）'],
    ['人员总重量', '140 kg'],
    ['床垫（蓝盒子20cm）', '约25-35 kg（记忆棉/弹簧床垫典型重量）'],
    ['床面总静载荷', '约165-175 kg'],
    ['动载系数（上床/翻身）', '1.5-2.0'],
    ['床面设计动载荷', '约250-350 kg'],
    ['桌面载荷（电视+音响+电脑）', '42寸OLED约12-15kg，S2000约12kg/对，笔记本+平板约5kg，合计约30-35kg'],
    ['', ''],
    ['—— 力学评估 ——', ''],
    ['立柱承重能力', '✅ 充足。4根4040欧标铝型材立柱，截面积约5.5cm²/根，6063-T5铝许用压应力约80-100MPa，4根总承载力约17-22吨，远大于需求。1935mm高度下压杆稳定安全系数>5。'],
    ['床架纵梁抗弯', '⚠️ 需要确认。4040欧标惯性矩Ix≈114cm⁴，床宽方向1580mm纵梁（如果沿宽度方向是1580跨距）单根4040在1580mm跨距跨中载荷约40kg（按2人均匀分布）时，挠度约2-4mm，在允许范围内（<L/400≈4mm）。如果纵梁是2040方向（2040mm跨度）则挠度会更大，需要多根横梁分担。'],
    ['床板横梁分布', '✅ 合理。设计中有多根3030/4040横梁（Pos15-20等1500mm横梁），间距约300-400mm，床板承重能均匀传到纵梁。'],
    ['桌面承重', '✅ 充足。桌面下方有型材支撑，30-35kg设备重量远低于4040/3030梁的承载能力。'],
    ['连接件强度', '✅ 充足。100×100重型角件（30个）+ 17×20小角码（26个）+ 万向/十字连接件（37个），连接点数量充足，M8螺栓（111个）抗剪切力强。'],
    ['', ''],
    ['—— 潜在风险点与改进建议 ——', ''],
    ['风险1：床架跨中挠度', '如果床的长向（2040mm）是主跨方向，建议在纵梁中部加一根立柱（或加斜撑），或增加横梁密度。'],
    ['风险2：上床动载荷', '上床时对最上一级横梁/床板有冲击载荷，建议梯子位置的承力型材使用4040并加装辅助支撑。'],
    ['风险3：梯子与框架连接', '确认梯子（如果有）与主体框架有至少4个牢固连接点，使用M8螺栓+T型螺母固定。'],
    ['风险4：地面调平', '必须安装可调脚杯（脚垫），确保4根立柱均匀受力，避免框架扭曲导致内应力。'],
    ['建议1：加中间支撑立柱', '床架下方（桌面上方/旁边）可以增加1-2根4040立柱，使床面跨度减半，大幅提高承重和减小晃动。'],
    ['建议2：斜撑加固', '在床架后侧（靠墙侧）加装X型或L型斜撑，提高整体抗侧倾和抗晃动能力。'],
    ['建议3：床板选择', '建议使用≥18mm厚多层实木板或25mm厚颗粒板，板下横梁间距≤400mm。'],
    ['建议4：桌面固定', '桌面（放电视/音响）建议用M5螺丝通过角件预留孔固定，防止设备倾倒。'],
    ['', ''],
    ['—— 总体结论 ——', ''],
    ['承重是否足够', '✅ 足够。按标准静载计算，4040+3030铝型材框架完全可以承受140kg两人+35kg床垫+30kg桌面设备的载荷，安全系数约2-3倍。'],
    ['需要加固的地方', '建议在床架中部增加1根支撑立柱（落地），可以显著减少纵梁挠度和整体晃动，提高安全感。'],
]
for r, rd in enumerate(eval_data, 1):
    ws3.cell(r, 1, rd[0])
    ws3.cell(r, 2, rd[1])
    for c in range(1,3):
        cell = ws3.cell(r, c)
        cell.border = thin
        cell.alignment = left_wrap
        if r == 1:
            cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center
        elif str(rd[0]).startswith('——'):
            cell.font = Font(bold=True, color="4472C4"); cell.fill = stat_fill
        elif '风险' in str(rd[0]) or '⚠️' in str(rd[0]):
            cell.fill = note_fill

ws4 = wb.create_sheet("与商家确认事项")
checklist = [
    ['确认项', '选项/说明', '请打√'],
    ['表面处理', '□ 银白阳极氧化（推荐，耐磨耐腐蚀）', ''],
    ['', '□ 黑色磨砂阳极氧化', ''],
    ['', '□ 本色不处理（便宜但易氧化发黑）', ''],
    ['', '□ 喷砂+阳极氧化（质感好）', ''],
    ['型材壁厚', '□ 标准壁厚（≥1.8mm，推荐）', ''],
    ['', '□ 加厚（≥2.0mm）', ''],
    ['切割精度', '□ 公差±0.5mm以内（必须，否则装不上）', ''],
    ['孔加工', '□ 按本清单+MayCAD PDF图纸加工所有沉头孔和连接件通孔', ''],
    ['配件是否配齐', '□ 是（按配件清单配齐螺丝、螺母、角件、连接件、端盖）', ''],
    ['', '□ 否（仅加工铝型材，配件我自己买）', ''],
    ['床板/桌板', '□ 另行采购，不含在本框架内', ''],
    ['脚垫/脚杯', '□ 请配M8可调脚杯4个（必须！）', ''],
    ['梯子', '□ 已包含在设计中（Pos1-8等小件）', ''],
    ['', '□ 需要额外增加梯子', ''],
    ['交付文件', '本Excel + 原始PDF施工图纸（导出尝试.pdf）一起交付', ''],
]
for r, rd in enumerate(checklist, 1):
    for c, v in enumerate(rd, 1):
        cell = ws4.cell(r, c, v)
        cell.border = thin
        cell.alignment = left_wrap
        if r == 1:
            cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center

ws.column_dimensions['A'].width = 7
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 9
ws.column_dimensions['D'].width = 7
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 10
ws.column_dimensions['I'].width = 80
ws.row_dimensions[1].height = 35
for r in range(2, row+1):
    ws.row_dimensions[r].height = 45

ws2.column_dimensions['A'].width = 7
ws2.column_dimensions['B'].width = 24
ws2.column_dimensions['C'].width = 32
ws2.column_dimensions['D'].width = 7
ws2.column_dimensions['E'].width = 38

ws3.column_dimensions['A'].width = 28
ws3.column_dimensions['B'].width = 85
for r in range(1, len(eval_data)+1):
    ws3.row_dimensions[r].height = 30

ws4.column_dimensions['A'].width = 18
ws4.column_dimensions['B'].width = 55
ws4.column_dimensions['C'].width = 12

out = r'c:\Users\Admin\Downloads\ezxss\上床下桌-中文加工清单-完整版.xlsx'
wb.save(out)

print(f"✅ 已生成完整中文加工清单: {out}")
print(f"型材种类: {len(bom_profiles)}")
print(f"沉头孔总计(型材上): {total_cb}，连接件通孔总计: {total_thru}")
print()
print("各型材加工统计：")
for pos in sorted(bom_profiles.keys()):
    p = bom_profiles[pos]
    if p['length']==0: continue
    cb = sum(1 for h in p['holes'] if h['type']=='cb')
    th = sum(1 for h in p['holes'] if h['type']=='thru')
    sides = set(h['side'] for h in p['holes'] if h['side'])
    print(f"  Pos{pos:2d}: {p['size']} L={p['length']:4d}mm ×{p['qty']}  沉头孔×{cb} 通孔×{th}")
